# coding=utf-8

"""
Data Pre-processing for PES (Pro Evolution Soccer) Dataset

This script focuses on loading and pre-processing player data from PES game files 
(specifically, XLSX files exported from the game). It's likely part of a larger 
project that analyzes or simulates football (soccer) performance based on this data.

Data Sources:
* Goalkeeper.xlsx: Contains information about goalkeepers (ratings, skills).
* Back.xlsx: Contains data for defenders.
* Forward.xlsx: Includes information about forwards (strikers and wingers).

Original Data Location:
https://github.com/ShenbaoYu/TCFPACN 

Note: This script requires the 'FBTP' module (presumably a custom module for football 
data analysis). Make sure it's installed and accessible within your project.
"""

import os, sys
BASE_DIR = os.path.dirname(os.path.dirname(os.getcwd())) 
sys.path.append(BASE_DIR)
sys.path.append('TCFPACN')  # Add the TCFPACN module's path to the search path

import openpyxl  # Library for working with Excel files
import math
from FBTP import players  # Custom module for handling player objects
from collections import OrderedDict  # For maintaining order of player abilities
import requests # To download the Excel files

# URLs of the Excel files to download
urls = [
    'https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Goalkeeper.xlsx',
    'https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Back.xlsx',
    'https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Forward.xlsx',
]

# Download each file and save it locally
for url in urls:
    filename = url.split('/')[-1]  
    response = requests.get(url)
    with open(filename, 'wb') as f:
        f.write(response.content)


def read_criteria(path, criteria_file):
    """
    Reads and normalizes player evaluation criteria from a text file.
    
    Args:
        path: The directory path where the criteria file is located.
        criteria_file: The name of the text file containing criteria.

    Returns:
        A dictionary where keys are criteria names and values are normalized weights.
    """
    criteria = {}
    with open(path + criteria_file, 'r') as cf:
        while True:
            line = cf.readline()
            if not line:
                break
            name, value = line.split(":")  # Split line into name and value
            criteria[name] = int(value) 

    criteria = normalize(criteria)  # Normalize criteria values (ensure they sum to 1)
    return criteria


def get_goalkeepers(file_name):
    """
    Extracts goalkeeper data from an Excel file and creates Goalkeeper objects.

    Args:
        file_name: The name of the XLSX file containing goalkeeper data.

    Returns:
        A list of Goalkeeper objects, each representing a goalkeeper with their 
        ID, rating, calculated salary, and skills.
    """
    wb = openpyxl.load_workbook(file_name) 
    ws = wb["GoalKeeper"]
    
    goal_keepers = []
    for row in range(2, ws.max_row + 1):  # Start from row 2 to skip headers
        gk_id = ws.cell(row=row, column=1).value
        gk = players.Goalkeeper(gk_id)  
        gk.rating = ws.cell(row=row, column=10).value
        gk.salary = 0.0006375 * math.exp(0.1029 * gk.rating)  # Calculate salary based on rating
        
        for column in range(29, ws.max_column + 1):  # Read skill values
            gk.ability.append(ws.cell(row=row, column=column).value)
        goal_keepers.append(gk)  

    return goal_keepers


def read_info(file_name):
    """
    Reads and organizes player information (non-goalkeepers) from an Excel file.

    Args:
        file_name: The name of the XLSX file containing player data.

    Returns:
        Several dictionaries containing:
            - ability_name_id: Mapping of ability names to IDs
            - player_attributes: Player attributes (team, nationality)
            - player_abilities_name: Player skills
            - player_position: Player positions
            - player_rating: Player ratings
            - player_no_id: Mapping of player numbers to IDs
    """
    
    ability_name_id = {}  # ability {name:id}
    player_attributes = {}  # playbers attributes, including club and nationality
    player_abilities_name = OrderedDict()  # personal abilities
    player_position = {}  # players' position {id:position}
    player_rating = {}  # players' rating {id:rating}

    wb = openpyxl.load_workbook(file_name)
    ws = wb["Players"]

    player_no_id = {}  # player's number : id
    no = 0  # number of player
    ability_id = 0  # number of ability

    for row in range(1, ws.max_row + 1):
        if row == 1:
            for column in range(11, 29):  # target ability
                player_abilities_name[ws.cell(row=row, column=column).value] = {}

            for ability_name in player_abilities_name.keys():
                ability_name_id[ability_name] = ability_id
                ability_id += 1
        else:
            player_id = ws.cell(row=row, column=1).value  # get player's ID
            player_no_id[no] = player_id
            position = ws.cell(row=row, column=2).value  # get player's position
            player_position[no] = position
            rating = ws.cell(row=row, column=10).value  # get player's rating
            player_rating[no] = rating

            # player's attributes, including club and nationality
            team_name = ws.cell(row=row, column=4).value  # player's team name
            nationality = ws.cell(row=row, column=5).value  # player's nationality
            player_attributes[no] = []
            player_attributes[no].append(team_name)
            player_attributes[no].append(nationality)
            
            # player's abilities
            _ = []
            for column in range(11, 29):  # range of target abilities
                _.append(ws.cell(row=row, column=column).value)
            c = 0
            for ability in player_abilities_name:
                player_abilities_name[ability][no] = _[c]
                c += 1
            
            no += 1

    return ability_name_id,       \
           player_attributes,     \
           player_abilities_name, \
           player_position,       \
           player_rating,         \
           player_no_id


def normalize(dict_type):
"""
- Helper function used by read_criteria to normalize the values of a dictionary.
- Divides each value by the total sum of the values, ensuring that the sum of the normalized values is 1.
"""
    total = sum(v for v in dict_type.values())
    tmp = {}
    for key, value in dict_type.items():
        tmp[key] = value / total
    return tmp
    
    
goalkeepers = get_goalkeepers('Goalkeeper.xlsx')
ability_name_id, player_attributes, player_abilities_name, player_position, player_rating, player_no_id = read_info('Back.xlsx')
ability_name_id, player_attributes, player_abilities_name, player_position, player_rating, player_no_id = read_info('Forward.xlsx')
